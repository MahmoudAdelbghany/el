import re
import openpyxl
header_file_path = 'headers.h'
excel_file_path = 'output.xlsx'
with open(header_file_path, 'r') as file:
    content = file.read()
#  regex pattern to match function prototypes (i asked copilot to generate it and it works.  cool)
pattern = r'[a-zA-Z_]\w*\s+\**\s*[a-zA-Z_]\w*\s*\([^)]*\)\s*;'
prototypes = re.findall(pattern, content)
workbook = openpyxl.Workbook()
sheet = workbook.active
row = 1
id_counter = 1
for prototype in prototypes:
    sheet.cell(row=row, column=1, value=prototype)
    sheet.cell(row=row, column=2, value=f'IDX{id_counter:03d}')
    row += 1
    id_counter += 1
workbook.save(excel_file_path)
print("Function prototypes have been successfully written to the Excel file.")
